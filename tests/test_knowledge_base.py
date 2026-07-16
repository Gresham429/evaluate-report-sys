"""基础表存储与版本管理测试。

指纹值均为三份真实 Excel 的实测值。

关于「改一份真 Excel 的系数」：基础表 I 列（修正系数）在三份素材里**都是公式**
（如 `=办公房地产比较法!X9`），真身在比较法表 X 列。openpyxl 存盘会丢弃全部公式
缓存值（ADR-001 已实证），故直接复制再存会让 28 个系数全塌成 0——那样「改系数
→ 指纹变」会因为无关原因变绿。压平法见 `_flat_copy`。
"""

import json
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from src.engine.knowledge import BASE_SHEET, Factor, Knowledge, extract_knowledge
from src.extractor.condition import read_survey_conditions
from src.knowledge_base import BaseTableStore, canonical_form, fingerprint
from src.model import Category
from tests.conftest import CASES

_COEFF_COL = 9  # 基础表 I 列
_SCORES = (2, 1, 0, -1, -2)


def _grouped_direct(case: str) -> Knowledge:
    """直接读 Excel 并按实勘表分组——镜像 `BaseTableStore.import_from_excel` 的
    分组转换，用作跟落库/取出结果比对的基准（分组：见 store.py 的导入落库处）。
    """
    knowledge = extract_knowledge(CASES[case])
    group_of = {c.factor: c.group for c in read_survey_conditions(CASES[case])}
    return Knowledge(
        factors=tuple(
            Factor(
                row=f.row,
                name=f.name,
                levels=f.levels,
                coefficient=f.coefficient,
                group=group_of.get(f.name, ""),
            )
            for f in knowledge.factors
        ),
        scores=knowledge.scores,
    )


def _flat_copy(case: str, dest: Path) -> Path:
    """把真 Excel 压平成值后另存：以 `data_only=True` 载入即把公式换成缓存值。

    压平本身是**指纹中性**的（`test_flat_copy_keeps_fingerprint` 守着这一点），
    故其后指纹一旦变化，只可能是被测的那次改动造成的。
    """
    shutil.copy(CASES[case], dest)
    workbook = openpyxl.load_workbook(dest, data_only=True)
    workbook.save(dest)
    return dest


def _copy_with_coefficient(case: str, dest: Path, row: int, value: float) -> Path:
    """压平一份真 Excel，并改掉基础表 I 列某一行的修正系数——模拟估价师调系数。"""
    path = _flat_copy(case, dest)
    workbook = openpyxl.load_workbook(path)
    workbook[BASE_SHEET].cell(row, _COEFF_COL).value = value
    workbook.save(path)
    return path


def _factor(row: int) -> Factor:
    return Factor(
        row=row,
        name="重要场所距离",
        levels={"距区域中心距离＜1KM": 2, "距区域中心距离＞2.5KM": -2},
        coefficient=1.0,
    )


# ---------------------------------------------------------------- 指纹口径


def test_flat_copy_keeps_fingerprint(tmp_path: Path) -> None:
    """守卫：压平不改变知识，故不改变指纹。

    本断言一旦红，说明「改系数」类测试的前提塌了——那些测试会因为压平丢了
    系数（而非因为改动）而变绿，等于没测。
    """
    for case in ("农用", "办公", "商业"):
        flat = _flat_copy(case, tmp_path / f"{case}.xlsx")
        assert extract_knowledge(flat) == extract_knowledge(CASES[case])
        assert fingerprint(extract_knowledge(flat)) == fingerprint(extract_knowledge(CASES[case]))


def test_fingerprints_differ_across_categories() -> None:
    """三份真基础表的指纹两两不同——三类的因素与系数本就完全不同。"""
    prints = {c: fingerprint(extract_knowledge(CASES[c])) for c in ("农用", "办公", "商业")}
    assert len(set(prints.values())) == 3, f"指纹撞车：{prints}"


@pytest.mark.parametrize(
    ("case", "expected"),
    [
        ("农用", "2392ea9d6359"),
        ("办公", "95d043e06567"),
        ("商业", "df20ba03d3fb"),
    ],
)
def test_fingerprints_of_real_excels(case: str, expected: str) -> None:
    """三份真基础表的实测指纹，已回填决策记录 §3.3。

    锁住它们是为了让「改口径」这件事发出声音：口径一改，这三个值全变，全部
    历史版本号即失效（旧文件名再也对不上重算出的指纹）。那时该红，红了才好
    做决定——是回滚口径，还是接受历史版本作废并更新文档。

    （决策记录 §3.3 原记的 56ce11aa5505/6584dc9bf9d2/5df34ffb0b64 是探索期临时
    脚本所算，拼接口径未留存，与本口径对不上，属预期之内——见计划 §3.1。）
    """
    assert fingerprint(extract_knowledge(CASES[case])) == expected


def test_fingerprint_format() -> None:
    """12 位十六进制小写——文件名要用它。"""
    got = fingerprint(extract_knowledge(CASES["办公"]))
    assert len(got) == 12
    assert all(c in "0123456789abcdef" for c in got)


@pytest.mark.parametrize("case", ["农用", "办公", "商业"])
def test_fingerprint_is_deterministic(case: str) -> None:
    """同一份 Excel 算十次必须同值——否则每次导入都成「新版本」。

    这条测不出多少东西（同进程内 dict 插入顺序必然相同，去掉 sort_keys 它照样
    绿），真正钉住确定性的是下面的 `test_fingerprint_ignores_level_order` 与
    `test_fingerprints_of_real_excels`。留着它是因为它把「导十次必须同值」这个
    要求写在了明面上。
    """
    got = {fingerprint(extract_knowledge(CASES[case])) for _ in range(10)}
    assert len(got) == 1


def test_fingerprint_ignores_level_order() -> None:
    """档次字典的插入顺序不得影响指纹——`sort_keys` 的守卫。

    这才是能红的那个：去掉 canonical_form 里的 sort_keys，本测立刻失败。
    档次是「描述→分值」的映射，其在 dict 里的先后不是知识。
    """
    forward = Factor(row=3, name="临街状况", levels={"四面临街": 2, "无临街": -2}, coefficient=1.0)
    reverse = Factor(row=3, name="临街状况", levels={"无临街": -2, "四面临街": 2}, coefficient=1.0)
    assert forward.levels != reverse.levels or list(forward.levels) != list(reverse.levels)
    assert fingerprint(Knowledge((forward,), _SCORES)) == fingerprint(
        Knowledge((reverse,), _SCORES)
    )


def test_canonical_form_is_stable_and_sorted() -> None:
    """规范串是指纹的唯一输入，也是「什么算估价知识」的落地处，钉一次字符串。

    它必须不含行号，且档次按键排序。
    """
    got = canonical_form(Knowledge((_factor(7),), _SCORES))
    assert got == (
        '{"分值标尺":[2,1,0,-1,-2],'
        '"因素":[{"名称":"重要场所距离",'
        '"档次":{"距区域中心距离＜1KM":2,"距区域中心距离＞2.5KM":-2},'
        '"系数":"1.0"}]}'
    )
    assert "行号" not in got and "7" not in got


def test_row_is_not_part_of_fingerprint() -> None:
    """只有行号不同 → 同一份知识 → 同一指纹。

    基础表中间插一个空行，全部因素行号后移，但估价知识一个字没变。
    """
    shifted = Knowledge(factors=(_factor(3),), scores=_SCORES)
    original = Knowledge(factors=(_factor(4),), scores=_SCORES)
    assert original != shifted, "row 是 Knowledge 的一部分（存取往返要用）"
    assert fingerprint(original) == fingerprint(shifted), "但 row 不是知识，不该进指纹"


def test_fingerprint_covers_coefficient() -> None:
    """系数变 → 指纹变。"""
    base = Knowledge(factors=(_factor(3),), scores=_SCORES)
    changed = Knowledge(
        factors=(Factor(row=3, name="重要场所距离", levels=_factor(3).levels, coefficient=5.0),),
        scores=_SCORES,
    )
    assert fingerprint(base) != fingerprint(changed)


def test_fingerprint_covers_level_score() -> None:
    """同一档次描述配的分值变 → 指纹变。"""
    base = Knowledge(factors=(_factor(3),), scores=_SCORES)
    changed = Knowledge(
        factors=(
            Factor(
                row=3,
                name="重要场所距离",
                levels={"距区域中心距离＜1KM": 1, "距区域中心距离＞2.5KM": -2},
                coefficient=1.0,
            ),
        ),
        scores=_SCORES,
    )
    assert fingerprint(base) != fingerprint(changed)


def test_fingerprint_covers_factor_name_and_scale() -> None:
    """因素名、分值标尺变 → 指纹变。"""
    base = Knowledge(factors=(_factor(3),), scores=_SCORES)
    renamed = Knowledge(
        factors=(Factor(row=3, name="别的因素", levels=_factor(3).levels, coefficient=1.0),),
        scores=_SCORES,
    )
    rescaled = Knowledge(factors=(_factor(3),), scores=(4, 2, 0, -2, -4))
    assert fingerprint(base) != fingerprint(renamed)
    assert fingerprint(base) != fingerprint(rescaled)


# ---------------------------------------------------------------- 存取往返


@pytest.mark.parametrize("case", ["农用", "办公", "商业"])
def test_roundtrip_is_lossless(tmp_path: Path, case: str) -> None:
    """存→取的 Knowledge 必须与直接读 Excel（含按实勘表分组）的完全相等，含 row。

    `direct` 用 `_grouped_direct` 而非裸 `extract_knowledge`：导入会按实勘表把
    factor.group 填上（见 store.py 的导入落库处），裸读不会，两者不再相等是
    分组功能本身的预期结果，不是回归。
    """
    store = BaseTableStore(tmp_path)
    result = store.import_from_excel(CASES[case], now=datetime(2026, 7, 16, 10, 0))
    direct = _grouped_direct(case)
    loaded = store.load(Category(case), result.版本.指纹)

    assert loaded == direct
    # row 单列出来盯：Excel 路径靠它定位比较法表的行（adapter.py 的 f.row + 6），
    # 丢了不会报错，只会静默读错行。
    assert [f.row for f in loaded.factors] == [f.row for f in direct.factors]
    assert loaded.scores == direct.scores
    assert fingerprint(loaded) == result.版本.指纹


def test_load_defaults_to_latest_import(tmp_path: Path) -> None:
    """不给指纹时取最新导入的一版——以台账的导入时间为准。"""
    store = BaseTableStore(tmp_path)
    old = store.import_from_excel(
        _copy_with_coefficient("办公", tmp_path / "旧.xlsx", 3, 1),
        now=datetime(2026, 3, 1, 9, 0),
    )
    new = store.import_from_excel(
        _copy_with_coefficient("办公", tmp_path / "新.xlsx", 3, 7),
        now=datetime(2026, 9, 1, 9, 0),
    )
    assert store.load(Category.OFFICE).factors[0].coefficient == 7.0
    assert store.load(Category.OFFICE, old.版本.指纹).factors[0].coefficient == 1.0
    assert store.load(Category.OFFICE, new.版本.指纹).factors[0].coefficient == 7.0


def test_latest_is_by_import_time_not_import_order(tmp_path: Path) -> None:
    """「最新」以台账的导入时间为准，不是文件名排序、也不是入库先后。

    故意让导入顺序与时间顺序相反：先导入时间较晚的那版。按文件名排序（指纹是
    哈希，字典序毫无意义）或按入库先后取，都会在此翻车。
    """
    store = BaseTableStore(tmp_path)
    later = store.import_from_excel(
        _copy_with_coefficient("办公", tmp_path / "晚.xlsx", 3, 7), now=datetime(2026, 9, 1)
    )
    store.import_from_excel(
        _copy_with_coefficient("办公", tmp_path / "早.xlsx", 3, 1), now=datetime(2026, 3, 1)
    )
    current = store.current(Category.OFFICE)
    assert current is not None
    assert current.指纹 == later.版本.指纹
    assert store.load(Category.OFFICE).factors[0].coefficient == 7.0


# ---------------------------------------------------------------- 版本管理


def test_changed_coefficient_creates_new_version(tmp_path: Path) -> None:
    """改一个修正系数 → 指纹必变、落成新文件、旧文件仍在、旧版本仍可取回重算。

    这是本模块存在的理由：2026-09 估价师调了系数，2027-01 复查 2026-03 那份报告，
    须能拿回 v1 把 2.83 重算出来（决策记录 §3.1）。
    """
    store = BaseTableStore(tmp_path)
    v1 = store.import_from_excel(
        _copy_with_coefficient("办公", tmp_path / "v1.xlsx", 3, 1), now=datetime(2026, 3, 1, 9, 0)
    )
    v2 = store.import_from_excel(
        _copy_with_coefficient("办公", tmp_path / "v2.xlsx", 3, 5), now=datetime(2026, 9, 1, 9, 0)
    )

    assert v1.是否新版 is True
    assert v2.是否新版 is True
    assert v1.版本.指纹 != v2.版本.指纹, "系数变了，指纹必须跟着变"

    # 旧版本永不覆盖——「能复现」是可复查的前提
    assert (tmp_path / f"办公-{v1.版本.指纹}.json").exists()
    assert (tmp_path / f"办公-{v2.版本.指纹}.json").exists()
    assert store.load(Category.OFFICE, v1.版本.指纹).factors[0].coefficient == 1.0
    assert store.load(Category.OFFICE, v2.版本.指纹).factors[0].coefficient == 5.0
    assert len(store.list_versions(Category.OFFICE)) == 2


def test_reimport_same_excel_is_not_a_new_version(tmp_path: Path) -> None:
    """同一份 Excel 导两次：第二次不是新版，不新增文件，台账不重复记。"""
    store = BaseTableStore(tmp_path)
    first = store.import_from_excel(CASES["办公"], now=datetime(2026, 7, 16, 10, 0))
    second = store.import_from_excel(CASES["办公"], now=datetime(2026, 7, 17, 11, 30))

    assert first.是否新版 is True
    assert second.是否新版 is False
    assert second.版本.指纹 == first.版本.指纹
    assert second.版本.导入时间 == first.版本.导入时间, "重导不得改写首次导入时间"
    assert len(list(tmp_path.glob("办公-*.json"))) == 1
    assert len(store.list_versions(Category.OFFICE)) == 1
    ledger = json.loads((tmp_path / "台账.json").read_text(encoding="utf-8"))
    assert len(ledger) == 1


def test_reimport_restores_a_deleted_version_file(tmp_path: Path) -> None:
    """版本文件被误删后，重导同一份 Excel 必须能把它救回来。

    手里还攥着源头 Excel 却救不回旧版本，直接打在「能复现」上；而重导同一份
    Excel 正是估价师唯一会想到的自救动作。指纹命中只查台账不查文件的话，这里
    会返回「已在库中」却不补文件，该版本从此永久取不出。
    """
    store = BaseTableStore(tmp_path)
    first = store.import_from_excel(CASES["办公"], now=datetime(2026, 3, 1, 9, 0))
    target = tmp_path / f"办公-{first.版本.指纹}.json"
    target.unlink()

    again = store.import_from_excel(CASES["办公"], now=datetime(2026, 9, 1, 9, 0))
    assert target.exists(), "文件没了却不补，该版本永久取不出"
    assert store.load(Category.OFFICE, first.版本.指纹) == _grouped_direct("办公")
    # 补文件不等于新版本：这版何时进的库是既成事实，不该被一次修复改写成今天。
    assert again.是否新版 is False
    assert again.版本.导入时间 == datetime(2026, 3, 1, 9, 0)
    assert len(json.loads((tmp_path / "台账.json").read_text(encoding="utf-8"))) == 1


def test_load_rejects_a_tampered_version_file(tmp_path: Path) -> None:
    """版本文件被手改过 → load 必须报错，不得静默返回改过的知识。

    文件是人类可读可手改的 JSON，手改迟早发生。而「拿 v1 重算 2026-03 那份报告」
    若静默用上改过的系数，算出的是另一个数且无人察觉——正是本模块要防的事故。
    """
    store = BaseTableStore(tmp_path)
    result = store.import_from_excel(CASES["办公"], now=datetime(2026, 7, 16, 10, 0))
    target = tmp_path / f"办公-{result.版本.指纹}.json"
    payload = json.loads(target.read_text(encoding="utf-8"))
    payload["因素"][0]["系数"] = 99.0  # 有人手痒改了个系数
    target.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    with pytest.raises(ValueError, match="疑被改动"):
        store.load(Category.OFFICE, result.版本.指纹)


def test_load_tolerates_cosmetic_edits(tmp_path: Path) -> None:
    """校验的是知识、不是字节：改缩进、调键序、把 1.0 写成 1 都不算改动。

    否则「人类可读可手改」这条就名存实亡——格式化一下文件就再也读不出来了。
    """
    store = BaseTableStore(tmp_path)
    result = store.import_from_excel(CASES["办公"], now=datetime(2026, 7, 16, 10, 0))
    target = tmp_path / f"办公-{result.版本.指纹}.json"
    payload = json.loads(target.read_text(encoding="utf-8"))
    payload["因素"][0]["系数"] = int(payload["因素"][0]["系数"])  # 1.0 → 1
    target.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")  # 压成一行

    assert store.load(Category.OFFICE, result.版本.指纹) == _grouped_direct("办公")


def test_reimport_does_not_rewrite_file(tmp_path: Path) -> None:
    """已存在的版本文件不得重写——不可变才谈得上可复现。

    往文件里塞个哨兵再重导：文件若被重写，哨兵就没了。比对 mtime 可靠——
    同秒内重写在部分文件系统上看不出时间差。
    """
    store = BaseTableStore(tmp_path)
    result = store.import_from_excel(CASES["办公"], now=datetime(2026, 7, 16, 10, 0))
    path = tmp_path / f"办公-{result.版本.指纹}.json"
    payload = json.loads(path.read_text(encoding="utf-8"))
    payload["哨兵"] = "不得被重写"
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    store.import_from_excel(CASES["办公"], now=datetime(2026, 7, 18, 10, 0))
    assert "哨兵" in json.loads(path.read_text(encoding="utf-8"))


def test_versions_of_different_categories_coexist(tmp_path: Path) -> None:
    """三类各自独立成版，互不影响。"""
    store = BaseTableStore(tmp_path)
    for case in ("农用", "办公", "商业"):
        store.import_from_excel(CASES[case], now=datetime(2026, 7, 16, 10, 0))
    for case in ("农用", "办公", "商业"):
        assert len(store.list_versions(Category(case))) == 1
        assert len(list(tmp_path.glob(f"{case}-*.json"))) == 1


def test_list_versions_newest_first(tmp_path: Path) -> None:
    """按导入时间新→旧。"""
    store = BaseTableStore(tmp_path)
    stamps = [datetime(2026, 3, 1, 9, 0), datetime(2026, 6, 1, 9, 0), datetime(2026, 9, 1, 9, 0)]
    for i, when in enumerate(stamps):
        store.import_from_excel(
            _copy_with_coefficient("办公", tmp_path / f"v{i}.xlsx", 3, i + 1), now=when
        )
    listed = store.list_versions(Category.OFFICE)
    assert [v.导入时间 for v in listed] == sorted(stamps, reverse=True)


def test_version_info_records_source_and_time(tmp_path: Path) -> None:
    """台账如实记录类别、指纹、导入时间、来源文件名。"""
    store = BaseTableStore(tmp_path)
    result = store.import_from_excel(CASES["办公"], now=datetime(2026, 7, 16, 10, 30))
    info = store.current(Category.OFFICE)
    assert info is not None
    assert info.类别 is Category.OFFICE
    assert info.指纹 == result.版本.指纹
    assert info.导入时间 == datetime(2026, 7, 16, 10, 30)
    assert info.来源文件名 == "办公实勘表、比较法.xlsx"

    entry = json.loads((tmp_path / "台账.json").read_text(encoding="utf-8"))[0]
    assert entry == {
        "类别": "办公",
        "指纹": result.版本.指纹,
        "导入时间": "2026-07-16T10:30:00",
        "来源文件名": "办公实勘表、比较法.xlsx",
    }


def test_import_time_defaults_to_now(tmp_path: Path) -> None:
    """不给 now 时取当前时间。（可注入性由上面固定时间的那些测试担着。）"""
    store = BaseTableStore(tmp_path)
    before = datetime.now()
    result = store.import_from_excel(CASES["办公"])
    assert before <= result.版本.导入时间 <= datetime.now()


# ---------------------------------------------------------------- 空库与错误


def test_current_is_none_when_nothing_imported(tmp_path: Path) -> None:
    store = BaseTableStore(tmp_path)
    assert store.current(Category.OFFICE) is None
    assert store.list_versions(Category.OFFICE) == ()


def test_load_without_any_version_raises_clearly(tmp_path: Path) -> None:
    """空库时报清楚的错，不得抛出 KeyError/IndexError 这类哑谜。"""
    store = BaseTableStore(tmp_path)
    with pytest.raises(FileNotFoundError, match="办公"):
        store.load(Category.OFFICE)


def test_load_unknown_fingerprint_raises_clearly(tmp_path: Path) -> None:
    store = BaseTableStore(tmp_path)
    store.import_from_excel(CASES["办公"], now=datetime(2026, 7, 16, 10, 0))
    with pytest.raises(FileNotFoundError, match="deadbeefcafe"):
        store.load(Category.OFFICE, "deadbeefcafe")


def test_missing_ledger_is_an_empty_store(tmp_path: Path) -> None:
    """目录不存在即空库，不是错误——首次运行时它本就不存在。"""
    store = BaseTableStore(tmp_path / "从未建过")
    assert store.current(Category.OFFICE) is None


# ---------------------------------------------------------------- 文件形态


def test_json_is_human_readable(tmp_path: Path) -> None:
    """版本文件与台账须人类可读可手改——与实例库一致。"""
    store = BaseTableStore(tmp_path)
    result = store.import_from_excel(CASES["办公"], now=datetime(2026, 7, 16, 10, 0))
    for path in (tmp_path / f"办公-{result.版本.指纹}.json", tmp_path / "台账.json"):
        text = path.read_text(encoding="utf-8")
        assert "办公" in text, "中文不得被转义成 \\uXXXX"
        assert "\n" in text, "须缩进换行，不得压成一行"
    body = (tmp_path / f"办公-{result.版本.指纹}.json").read_text(encoding="utf-8")
    assert "重要场所距离" in body
