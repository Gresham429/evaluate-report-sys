"""年租赁价值算术。

断言值全部来自三份金样 Excel 的一览表实测（L 列 = 公式算出的缓存值），
不是自己编的期望值。
"""

import openpyxl
import pytest

from src.engine.annual import DAYS_PER_YEAR, annual_value
from src.model import Category
from tests.conftest import CASES

# 实测：一览表 L 列公式与算出的值。
#   办公 L49 =ROUND(J49*K49*365,0) → 368030      （J49=356.29 K49=2.83）
#   农用 L49 =K49*J49              → 70000       （J49=50     K49=1400）
#   商业 L49 =ROUND(J49*K49*365,0) → 72708       （J49=60     K49=3.32）
GOLDEN_ROWS = [
    (Category.OFFICE, 356.29, 2.83, 368030),
    (Category.OFFICE, 367.4, 2.83, 379506),
    (Category.AGRICULTURAL, 50.0, 1400.0, 70000),
    (Category.COMMERCIAL, 60.0, 3.32, 72708),
    (Category.COMMERCIAL, 70.0, 3.32, 84826),
]


@pytest.mark.parametrize(("category", "area", "unit_price", "expected"), GOLDEN_ROWS)
def test_reproduces_golden_rows(
    category: Category, area: float, unit_price: float, expected: int
) -> None:
    """三份金样一览表的每一行都必须算得出来。"""
    assert annual_value(category, area, unit_price) == expected


def test_agricultural_does_not_multiply_by_days() -> None:
    """农用按年计租，房屋类按天计租——两者差 365 倍，混了即错 365 倍。"""
    land = annual_value(Category.AGRICULTURAL, 50.0, 1400.0)
    house = annual_value(Category.OFFICE, 50.0, 1400.0)
    assert land == 70000
    assert house == 70000 * DAYS_PER_YEAR


def test_rounds_half_up_like_excel() -> None:
    """Excel 的 ROUND 逢五进一；Python 内置 round() 是银行家舍入，会舍成偶数。

    1㎡ × 0.5元/㎡·天 × 365 = 182.5：Excel 记 183，内置 round() 记 182。
    """
    assert annual_value(Category.OFFICE, 1.0, 0.5) == 183
    assert annual_value(Category.AGRICULTURAL, 1.0, 0.5) == 1


def test_zero_area_or_price_is_zero() -> None:
    """空行/未填单价不该炸，算 0 即可——是否为问题交给校验器提示。"""
    assert annual_value(Category.OFFICE, 0.0, 2.83) == 0
    assert annual_value(Category.OFFICE, 356.29, 0.0) == 0


@pytest.mark.parametrize("case", ["农用", "办公", "商业"])
def test_matches_every_row_of_real_excel(case: str) -> None:
    """直接拿三份真 Excel 的一览表逐行对——防止上面的常量抄错。"""
    category = Category(case if case != "农用" else "农用")
    sheet_name = (
        "办公房地产实地查勘记录表" if category is Category.OFFICE
        else "房地产实地查勘记录表"
    )
    sheet = openpyxl.load_workbook(CASES[case], data_only=True)[sheet_name]
    checked = 0
    row = 49
    while isinstance(sheet.cell(row, 6).value, int | float):
        area = float(sheet.cell(row, 10).value)
        unit_price = float(sheet.cell(row, 11).value)
        excel_value = int(sheet.cell(row, 12).value)
        assert annual_value(category, area, unit_price) == excel_value
        checked += 1
        row += 1
    assert checked > 0, f"{case} 一览表没读到任何行"
