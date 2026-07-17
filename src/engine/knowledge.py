"""从 Excel 基础表提取比较法知识。

新 C1：基础表的知识（因素、档次、系数、分值）是既定正确的，系统只读不改。
用户改系数/档次/因素 → 改 Excel 基础表即生效，本模块不需改动。

实测：三类的修正系数完全不同（农用 [2,1,2,1,2,0]、办公 [1,1,2,2,1,3]、
商业 [2,2,2,0,1,2]），故必须按上传的 Excel 逐份读取，不存在全公司主表。
"""

import logging
from collections.abc import Mapping
from dataclasses import dataclass
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

__all__ = [
    "Factor",
    "Knowledge",
    "extract_knowledge",
    "BASE_SHEET",
    "parse_range",
    "apply_coefficient_overrides",
]

BASE_SHEET = "比较因素条件说明表（基础表）"

_FIRST_FACTOR_ROW = 3
_LAST_FACTOR_ROW = 30
_SCORE_ROW = 31          # 分值标尺行：D..H = 2/1/0/-1/-2
_LEVEL_COLS = range(4, 9)  # D..H
_COEFF_COL = 9           # I 列：每差 1 档修正系数
_RANGE_COL = 10          # J 列：调整范围，如 "2-4"；估价师逐份调系数时的边界提示
_NAME_COLS = (3, 2)      # C 列优先（因素名），退到 B 列（分组名）


@dataclass(frozen=True)
class Factor:
    """一个比较因素。"""

    row: int
    name: str
    levels: dict[str, int]
    coefficient: float
    # 资产状况分组：区位状况/实物状况/权益状况。来自实勘表 A 列，导入时填。
    # 仅供表单/报告分组展示，**不进算术、不进指纹**（fingerprint.canonical_form 不含它）。
    group: str = ""
    # 单份报告调系数时的边界提示，来自基础表 J 列原文（如 "2-4"）。软提示、不硬卡：
    # 实测 I 列系数有时落在 J 范围外，故本字段仅供表单展示参考区间，不作校验依据。
    # 同 group：**不进算术、不进指纹**（canonical_form 不含它）。空串表示该行无范围。
    调整范围: str = ""


@dataclass(frozen=True)
class Knowledge:
    """一份基础表承载的全部比较法知识。"""

    factors: tuple[Factor, ...]
    scores: tuple[int, ...]


def _cell_text(sheet: object, row: int, col: int) -> str:
    value = sheet.cell(row, col).value  # type: ignore[attr-defined]
    return str(value).strip() if value is not None else ""


def extract_knowledge(path: Path) -> Knowledge:
    """读 Excel 基础表 → 比较法知识。

    Args:
        path: 实勘表 Excel 路径（内含基础表工作表）。

    Returns:
        Knowledge。因素按基础表行序排列，空行跳过。

    Raises:
        ValueError: 基础表缺失，或分值行不是 5 个整数。
    """
    workbook = openpyxl.load_workbook(path, data_only=True)
    if BASE_SHEET not in workbook.sheetnames:
        raise ValueError(f"工作簿缺少基础表 {BASE_SHEET!r}：{path}")
    sheet = workbook[BASE_SHEET]

    raw_scores = [sheet.cell(_SCORE_ROW, col).value for col in _LEVEL_COLS]
    if not all(isinstance(s, int) for s in raw_scores) or len(raw_scores) != 5:
        raise ValueError(
            f"基础表第 {_SCORE_ROW} 行应为 5 个整数分值，实为 {raw_scores!r}：{path}"
        )
    scores = tuple(int(s) for s in raw_scores)  # type: ignore[arg-type]

    factors: list[Factor] = []
    for row in range(_FIRST_FACTOR_ROW, _LAST_FACTOR_ROW + 1):
        name = next((t for c in _NAME_COLS if (t := _cell_text(sheet, row, c))), "")
        levels = {
            text: scores[i]
            for i, col in enumerate(_LEVEL_COLS)
            if (text := _cell_text(sheet, row, col))
        }
        if not name:
            # 无因素名的行不是真实因素——实测农用基础表第 23 行即为例证：
            # B/C 列均为空，但 D..H 遗留了与表头行（第 2 行）相同的通用占位文本
            # "好/较好/一般/较差/差"，系数为 0，属模板残留而非真实评估因素。
            continue
        coeff = sheet.cell(row, _COEFF_COL).value
        factors.append(
            Factor(
                row=row,
                name=name,
                levels=levels,
                coefficient=float(coeff) if isinstance(coeff, (int, float)) else 0.0,
                调整范围=_cell_text(sheet, row, _RANGE_COL),
            )
        )
    logger.debug("提取基础表 %s：%d 个因素", path.name, len(factors))
    return Knowledge(factors=tuple(factors), scores=scores)


def parse_range(text: str) -> tuple[float, float] | None:
    """解析「调整范围」原文，如 "2-4" → (2.0, 4.0)。

    只认「数字-数字」这一种形状：素材里的真实值全是这种（"1-2"/"2-4"/"3-10" 等）。
    空白、纯数字（如系数行的 "0"）、多段或非数字一律视为「无范围」而非报错——
    J 列并非每行都是范围（分值标尺行、部分空行），调用方应当能容忍地跳过它们，
    而不是让一行脏数据打断整份基础表的读取。

    Args:
        text: 原始文本，通常来自 Factor.调整范围。

    Returns:
        (下限, 上限)；文本为空或不合「数字-数字」形状时为 None。
    """
    stripped = text.strip()
    if not stripped:
        return None
    parts = stripped.split("-")
    if len(parts) != 2:
        return None
    try:
        low, high = float(parts[0]), float(parts[1])
    except ValueError:
        return None
    return (low, high)


def apply_coefficient_overrides(
    knowledge: Knowledge, overrides: Mapping[str, float]
) -> Knowledge:
    """按估价师逐份填写的系数覆盖 Knowledge——单份报告的「实际知识」。

    不改基础表本身（新 C1 仍然成立，基础表只读不改）：本函数只产出一份新
    Knowledge，供当次报告计算与台账落盘用；传入的 knowledge 不受影响、不被
    改写。未知因素名选择**报错**而非静默忽略——overrides 多半来自表单/前端，
    键入错误理应在这里被截住，否则估价师会以为调整已生效，实则系数原封未动。

    Args:
        knowledge: 基础 Knowledge（通常来自 BaseTableStore.load）。
        overrides: {因素名: 新系数}。未出现在 overrides 里的因素系数不变。

    Returns:
        新 Knowledge：仅被 overrides 命中的因素的 coefficient 被替换，其余字段
        （row/levels/group/调整范围）与未命中的因素原样保留。

    Raises:
        ValueError: overrides 里出现了 knowledge 中不存在的因素名。
    """
    known_names = {f.name for f in knowledge.factors}
    unknown = sorted(set(overrides) - known_names)
    if unknown:
        raise ValueError(f"覆盖里出现未知因素名：{unknown}")
    factors = tuple(
        f
        if f.name not in overrides
        else Factor(
            row=f.row,
            name=f.name,
            levels=f.levels,
            coefficient=float(overrides[f.name]),
            group=f.group,
            调整范围=f.调整范围,
        )
        for f in knowledge.factors
    )
    return Knowledge(factors=factors, scores=knowledge.scores)
