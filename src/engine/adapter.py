"""从比较法表读取估价对象档次与三条实例。

依 C1：只读 `data_only=True` 缓存值，不重算、不回写。

实测行位：3=位置、4=成交价、5=面积、6=用途、7=交易情况、8=交易日期、
9..36=28 个因素槽位。列：D=估价对象、E/F/G=实例A/B/C；
指数列 L=基准、M/N/O=实例A/B/C。
"""

import logging
from pathlib import Path

import openpyxl

from src.engine.knowledge import extract_knowledge
from src.engine.methods.base import Instance
from src.extractor.field_map import comparison_sheet_name
from src.model import Category

logger = logging.getLogger(__name__)

__all__ = ["read_subject_levels", "read_instances", "SUBJECT_COL", "INSTANCE_COLS"]

SUBJECT_COL = 4                    # D
INSTANCE_COLS = (5, 6, 7)          # E/F/G
_INDEX_COLS = {5: 13, 6: 14, 7: 15}  # E→M, F→N, G→O
_BASE_INDEX_COL = 12               # L
_LOCATION_ROW = 3
_PRICE_ROW = 4
_TRADE_ROW = 7
_MARKET_ROW = 8
_FACTOR_OFFSET = 6                 # 基础表 3 行 ↔ 比较法 9 行


def _text(sheet: object, row: int, col: int) -> str:
    value = sheet.cell(row, col).value  # type: ignore[attr-defined]
    return str(value).strip() if value is not None else ""


def _normalize_market(subject_market: float, instance_market: float) -> float:
    """把 Excel 的「交易日期/市场状况修正」归一到引擎的 100 基约定（term = 指数/100）。

    两族 Excel 的这一项约定不同，本函数忠实复现各自的算法（见 docs/README §5）：

    - 现有三类（农用/办公/商业）：估价对象 row-8 恒为 **100**（100 基状况指数），
      Excel 修正 = 实例/对象 = `M8/L8` = `M8/100` → 直接用实例值，引擎 `实例/100` 复现。
    - 新四类（住宅/工业/停车场/建设用地）：row-8 是**原始交易日期指数**（如 3.24），
      Excel 修正 = 对象/实例 = `L8/M8` → 归一化 100 基值 = `100 × 对象/实例`，
      使引擎 `归一值/100 = L8/M8` 精确复现。

    ⚠️ **新四类的方向按当前提供的构造样例照搬，待真实案例 + 执业估价师复核**
    （4 份新 Excel 的 row-8 值完全相同，疑似占位数据）。这条是 Approach A 下
    「拿真实案例重锁算术金样」要复核的第一条。
    """
    if subject_market == 100.0:  # 100 基状况约定：估价对象恒为 100
        return instance_market
    if instance_market == 0.0:  # 防 0 除：实例日期指数缺失时按中性处理
        return subject_market
    return 100.0 * subject_market / instance_market


def read_subject_levels(path: Path, category: Category) -> dict[str, str]:
    """读估价对象的因素档次（比较法表 D 列）。

    Args:
        path: 实勘表 Excel 路径。
        category: 类别，决定工作表名。

    Returns:
        因素名 → 档次描述。
    """
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[comparison_sheet_name(category)]
    knowledge = extract_knowledge(path)
    return {
        f.name: _text(sheet, f.row + _FACTOR_OFFSET, SUBJECT_COL) for f in knowledge.factors
    }


def read_instances(path: Path, category: Category) -> tuple[Instance, ...]:
    """读三条实例（比较法表 E/F/G 列）。

    市场状况指数取自第 8 行的 M/N/O —— 那是估价师手填的常数，非公式。

    Args:
        path: 实勘表 Excel 路径。
        category: 类别，决定工作表名。

    Returns:
        三条 Instance。
    """
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[comparison_sheet_name(category)]
    knowledge = extract_knowledge(path)
    base_raw = sheet.cell(_MARKET_ROW, _BASE_INDEX_COL).value
    base = float(base_raw) if isinstance(base_raw, (int, float)) else 100.0

    instances: list[Instance] = []
    for col in INSTANCE_COLS:
        idx_col = _INDEX_COLS[col]
        price = sheet.cell(_PRICE_ROW, col).value
        trade = sheet.cell(_TRADE_ROW, idx_col).value
        market = sheet.cell(_MARKET_ROW, idx_col).value
        market_val = float(market) if isinstance(market, (int, float)) else base
        instances.append(
            Instance(
                位置=_text(sheet, _LOCATION_ROW, col),
                成交价=float(price) if isinstance(price, (int, float)) else 0.0,
                交易情况指数=float(trade) if isinstance(trade, (int, float)) else 100.0,
                市场状况指数=_normalize_market(base, market_val),
                因素档次={
                    f.name: _text(sheet, f.row + _FACTOR_OFFSET, col) for f in knowledge.factors
                },
            )
        )
    logger.debug("读取 %s 的 %d 条实例", path.name, len(instances))
    return tuple(instances)
