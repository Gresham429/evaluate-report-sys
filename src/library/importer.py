"""从现有实勘表 Excel 批量导入实例。

用户确认：批量导入的源就是现有的实勘表 Excel。实测三份素材可抽出 9 条。

抽取但不入库——须经估价师确认后再存，故本模块只返回不写盘。
"""

import logging
from pathlib import Path

import openpyxl

from src.engine.adapter import INSTANCE_COLS, read_instances
from src.engine.knowledge import extract_knowledge
from src.extractor.field_map import comparison_sheet_name, detect_category
from src.library.model import StoredInstance, make_id, parse_lease_start

logger = logging.getLogger(__name__)

__all__ = ["import_from_excel"]

_AREA_ROW = 5
_USAGE_ROW = 6
_TRADE_ROW = 7
_LEASE_ROW = 8


def import_from_excel(path: Path) -> tuple[StoredInstance, ...]:
    """从一份实勘表 Excel 抽取全部比较实例。

    Args:
        path: 实勘表 Excel 路径。

    Returns:
        抽取到的实例（**未入库**，供确认后再存）。

    Raises:
        ValueError: 类别无法识别或工作表缺失。
    """
    workbook = openpyxl.load_workbook(path, data_only=True)
    title = ""
    for name in workbook.sheetnames:
        if "实地查勘记录表" in name:
            title = str(workbook[name]["A1"].value or "")
            break
    category = detect_category(title)
    sheet = workbook[comparison_sheet_name(category)]
    knowledge = extract_knowledge(path)
    engine_instances = read_instances(path, category)

    result: list[StoredInstance] = []
    for engine_inst, col in zip(engine_instances, INSTANCE_COLS, strict=True):
        raw_lease = sheet.cell(_LEASE_ROW, col).value
        start, precision = parse_lease_start(raw_lease)
        area = sheet.cell(_AREA_ROW, col).value
        result.append(
            StoredInstance(
                编号=make_id(category, start, precision, engine_inst.位置),
                类别=category,
                位置=engine_inst.位置,
                成交价=engine_inst.成交价,
                面积=float(area) if isinstance(area, (int, float)) else 0.0,
                出租用途=str(sheet.cell(_USAGE_ROW, col).value or "").strip(),
                交易情况=str(sheet.cell(_TRADE_ROW, col).value or "").strip(),
                交易情况指数=engine_inst.交易情况指数,
                租期原文=str(raw_lease) if raw_lease is not None else "",
                起始日=start,
                日期精度=precision,
                因素档次=dict(engine_inst.因素档次),
            )
        )
    logger.info("从 %s 抽取 %d 条实例（共 %d 个因素）", path.name, len(result), len(knowledge.factors))
    return tuple(result)
