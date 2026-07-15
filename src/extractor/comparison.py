"""比较法表与估价结果一览表提取。

依约束 C1：只读缓存值。一览表的单价可能与比较法 T39 不同
（估价师手工取整，如农用 T39=1399.26 而 K49=1400），必须原样读取。
"""

import logging
from pathlib import Path

import openpyxl

from src.extractor.field_map import (
    COMPARISON_FIELDS,
    RESULT_COLUMNS,
    RESULT_FIRST_ROW,
    comparison_sheet_name,
    survey_sheet_name,
)
from src.model import Category, Subject

logger = logging.getLogger(__name__)

__all__ = ["extract_comparison", "extract_subjects"]


def _as_float(value: object, default: float = 0.0) -> float:
    """Excel 单元格值 → float。非数值（含 bool、日期、文本）一律取默认值。"""
    if isinstance(value, bool):
        return default
    if isinstance(value, (int, float)):
        return float(value)
    return default


def _as_int(value: object, default: int = 0) -> int:
    """Excel 单元格值 → int。非数值（含 bool、日期、文本）一律取默认值。"""
    if isinstance(value, bool):
        return default
    if isinstance(value, (int, float)):
        return int(value)
    return default


def _as_text(value: object) -> str:
    """Excel 单元格值 → 去空白的字符串。None 转空串。"""
    return str(value).strip() if value is not None else ""


def extract_comparison(path: Path, category: Category) -> dict[str, float]:
    """读比较法表的评估结果与离散度。

    Args:
        path: xlsx 路径。
        category: 类别，决定工作表名。

    Returns:
        `{"unit_price": T39, "dispersion": X37}`。

    Raises:
        ValueError: 工作表缺失。
    """
    workbook = openpyxl.load_workbook(path, data_only=True)
    name = comparison_sheet_name(category)
    if name not in workbook.sheetnames:
        raise ValueError(f"工作簿缺少工作表 {name!r}：{path}")
    sheet = workbook[name]
    result: dict[str, float] = {}
    for field, cell in COMPARISON_FIELDS.items():
        value = sheet[cell].value
        result[field] = _as_float(value)
    return result


def extract_subjects(path: Path, category: Category) -> tuple[Subject, ...]:
    """读估价结果一览表。

    从 RESULT_FIRST_ROW 起向下扫描，F 列序号不再是数字即停止。
    对象数可变（实测农用 1、办公 2、商业 2）。

    Args:
        path: xlsx 路径。
        category: 类别，决定工作表名。

    Returns:
        Subject 元组，按表中顺序。
    """
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[survey_sheet_name(category)]
    subjects: list[Subject] = []
    row = RESULT_FIRST_ROW
    while True:
        index = sheet.cell(row, RESULT_COLUMNS[0]).value
        if not isinstance(index, (int, float)):
            break
        values = [sheet.cell(row, col).value for col in RESULT_COLUMNS]
        subjects.append(
            Subject(
                index=_as_int(values[0]),
                owner=_as_text(values[1]),
                address=_as_text(values[2]),
                usage=_as_text(values[3]),
                area=_as_float(values[4]),
                unit_price=_as_float(values[5]),
                annual_value=_as_int(values[6]),
            )
        )
        row += 1
    logger.debug("提取一览表 %s：%d 个估价对象", path.name, len(subjects))
    return tuple(subjects)
