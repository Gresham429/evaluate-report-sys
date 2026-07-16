"""读实勘表 sheet 的逐因素〔组, 因素, 描述〕。

分组（区位/实物/权益）与逐因素手写描述都只在这里读一次，供两处复用：
基础表导入据此给因素分组、load_project 据此组装报告的资产状况三张表。

结构（三类真样本已核）：实勘表 sheet 名含「查勘记录表」；A 列在每组头一行
给出组名（办公写「权益状况(二)」，按前缀归一）；B 列因素名、D 列手写描述。
"""

import logging
from dataclasses import dataclass
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

__all__ = ["SurveyCondition", "read_survey_conditions", "GROUP_PREFIXES"]

SURVEY_SHEET_KEYWORD = "查勘记录表"
GROUP_PREFIXES = ("区位状况", "实物状况", "权益状况")
_COL_GROUP, _COL_FACTOR, _COL_DESC = 1, 2, 4  # A / B / D


@dataclass(frozen=True)
class SurveyCondition:
    group: str        # 区位状况 / 实物状况 / 权益状况（已按前缀归一）
    factor: str       # 因素名（B 列）
    description: str   # 手写描述（D 列），可空


def _text(sheet: object, row: int, col: int) -> str:
    v = sheet.cell(row, col).value  # type: ignore[attr-defined]
    return str(v).strip() if v is not None else ""


def _survey_sheet(workbook: openpyxl.Workbook) -> object:
    for name in workbook.sheetnames:
        if SURVEY_SHEET_KEYWORD in name:
            return workbook[name]
    raise ValueError(f"工作簿缺少实勘表（sheet 名含「{SURVEY_SHEET_KEYWORD}」）")


def _normalise_group(a_text: str) -> str | None:
    """A 列文字 → 归一化组名；非组标记返回 None。"""
    for prefix in GROUP_PREFIXES:
        if a_text.startswith(prefix):
            return prefix
    return None


def read_survey_conditions(path: Path) -> tuple[SurveyCondition, ...]:
    """读实勘表逐因素〔组, 因素, 描述〕，按出现顺序返回。

    Args:
        path: 实勘表 Excel 路径。

    Returns:
        每个有因素名的行一条；描述取 D 列（可空）。组按最近一次出现的组标记。

    Raises:
        ValueError: 找不到实勘表 sheet。
    """
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = _survey_sheet(workbook)
    out: list[SurveyCondition] = []
    current: str | None = None
    for row in range(1, sheet.max_row + 1):  # type: ignore[attr-defined]
        if (group := _normalise_group(_text(sheet, row, _COL_GROUP))) is not None:
            current = group
        factor = _text(sheet, row, _COL_FACTOR)
        if current is None or not factor:
            continue
        out.append(SurveyCondition(current, factor, _text(sheet, row, _COL_DESC)))
    logger.debug("实勘表 %s 读到 %d 条逐因素描述", path.name, len(out))
    return tuple(out)
