"""实勘表提取。

依约束 C1：只读 Excel 的计算缓存值（data_only=True），不重算任何数值。
"""

import logging
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

from src.extractor.field_map import SURVEY_FIELDS, detect_category, survey_sheet_name

logger = logging.getLogger(__name__)

__all__ = ["excel_serial_to_date", "extract_survey"]

# Excel 1900 日期系统的纪元（含 1900 闰年 bug 的补偿）
_EXCEL_EPOCH = date(1899, 12, 30)

# 需从序列号转为日期的字段
_DATE_FIELDS = frozenset({"survey_date", "value_date", "issue_date"})


def excel_serial_to_date(serial: float) -> date:
    """Excel 日期序列号 → date。

    Args:
        serial: Excel 存储的日期序列号，如 46132。

    Returns:
        对应日期，如 date(2026, 4, 20)。
    """
    return _EXCEL_EPOCH + timedelta(days=int(serial))


def _normalise_date(value: object) -> object:
    """把日期序列号或 datetime 统一成 YYYY-MM-DD 字符串。"""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return excel_serial_to_date(value).isoformat()
    return value


def extract_survey(path: Path) -> dict[str, object]:
    """读实勘表 → 字段字典。

    Args:
        path: xlsx 文件路径。

    Returns:
        含 `category` 及 SURVEY_FIELDS 全部键的字典。日期字段为 YYYY-MM-DD 字符串。

    Raises:
        ValueError: A1 标题无法识别类别，或工作表缺失。
    """
    workbook = openpyxl.load_workbook(path, data_only=True)
    # 先用任一实勘表的 A1 判类别：三类工作表名不同，故逐个试
    title = ""
    for name in workbook.sheetnames:
        if "实地查勘记录表" in name:
            title = str(workbook[name]["A1"].value or "")
            break
    category = detect_category(title)

    sheet_name = survey_sheet_name(category)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"工作簿缺少工作表 {sheet_name!r}：{path}")
    sheet = workbook[sheet_name]

    data: dict[str, object] = {"category": category}
    for field, cell in SURVEY_FIELDS.items():
        value = sheet[cell].value
        if field in _DATE_FIELDS:
            value = _normalise_date(value)
        elif value is not None:
            value = str(value).strip()
        data[field] = value
    logger.debug("提取实勘表 %s：类别=%s 报告号=%s", path.name, category, data["report_no"])
    return data
