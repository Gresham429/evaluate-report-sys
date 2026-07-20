"""从 7 份实勘表 xlsx 抽取每类别逐因素清单 + 档次选项 → 生成 miniprogram/factors.js。

- 因素名与分组：实勘表左列（区位/实物/权益状况）逐因素。
- 档次选项：同工作簿「比较因素条件说明表（基础表）」——每因素一行，D-H 列是
  好/较好/一般/较差/差 五档的具体口径（如「距区域中心距离＜5KM」），空档跳过。
现场采集页据此每因素渲染「描述(自由文字)+档次(下拉)」；因素名同时是
content.asset_conditions/subject_levels 的键（对齐办公端 backend.response_to_fields）。
因素/档次随 xlsx 变动就重跑本脚本；生成物勿手改。

用法：uv run python tools/gen_survey_factors.py
"""

from __future__ import annotations

import glob
import json
import os
import re

import openpyxl

# 文件名前缀 → 小程序类别值（对齐 form.js CATEGORIES / src.model.Category 值）
_FILE_TO_CATEGORY = {
    "住宅": "住宅",
    "停车场用地": "停车场用地",
    "农用地": "农用",
    "办公": "办公",
    "商业": "商业",
    "工业": "工业",
    "建设用地": "建设用地",
}

_SRC_GLOB = "../输入7种excel案例+输出两种报告模版/*实勘表、比较法.xlsx"
_OUT = "miniprogram/factors.js"


def _norm_section(text: str) -> str:
    """区位状况/实物状况/权益状况(二) → 去掉「(二)」等序号后缀。"""
    return re.sub(r"[(（].*?[)）]", "", text).strip()


def _cell(ws: object, r: int, c: int) -> str:
    v = ws.cell(r, c).value  # type: ignore[attr-defined]
    return str(v).strip() if v is not None and str(v).strip() else ""


def _extract_levels(wb: object) -> dict[str, list[str]]:
    """「比较因素条件说明表（基础表）」→ {因素名: [档次口径...]}。D-H 五档，空档跳过。"""
    name = next(
        (s for s in wb.sheetnames if "条件说明" in s or "比较因素条件" in s), None  # type: ignore[attr-defined]
    )
    if not name:
        return {}
    ws = wb[name]  # type: ignore[index]
    out: dict[str, list[str]] = {}
    # 只跳表头/脚注与大类名——注意「临街状况/临路状况」是因素、不能按 endswith("状况") 误杀。
    skip = {"级别", "修正系数", "资产状况", "区位状况", "实物状况", "权益状况"}
    for r in range(3, ws.max_row + 1):  # type: ignore[attr-defined]
        factor = _cell(ws, r, 3) or _cell(ws, r, 2)  # C 优先（区位），否则 B（实物/权益）
        if not factor or factor in skip or factor.isdigit():
            continue
        levels = [_cell(ws, r, c) for c in range(4, 9)]  # D,E,F,G,H
        levels = [x for x in levels if x]
        if levels and factor not in out:
            out[factor] = levels
    return out


def _extract(path: str) -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = next((s for s in wb.sheetnames if "查勘记录" in s or "实地查勘" in s), wb.sheetnames[0])
    ws = wb[sheet]
    levels_by_name = _extract_levels(wb)
    groups: list[dict[str, object]] = []
    index: dict[str, list[dict[str, object]]] = {}
    seen: dict[str, set[str]] = {}
    section = ""
    missing: list[str] = []
    for r in range(17, 46):  # 17 起是资产状况块；46 行是「现场查勘记录人员」不算因素
        a = _cell(ws, r, 1)
        name = _cell(ws, r, 2) or _cell(ws, r, 3)
        if a and "状况" in a and not a.startswith("资产"):
            section = _norm_section(a)
            if section not in index:
                index[section] = []
                seen[section] = set()
                groups.append({"section": section, "items": index[section]})
        if name and not name.isdigit() and section and name not in seen[section]:
            seen[section].add(name)
            levels = levels_by_name.get(name, [])
            if not levels:
                missing.append(name)
            index[section].append({"name": name, "levels": levels})
    if missing:
        print(f"  ⚠ 无匹配档次的因素（描述仍可填，档次下拉为空）：{missing}")
    return groups


def main() -> None:
    files = sorted(glob.glob(_SRC_GLOB))
    if not files:
        raise SystemExit(f"找不到实勘表：{_SRC_GLOB}")
    out: dict[str, list[dict[str, object]]] = {}
    for path in files:
        base = os.path.basename(path)
        prefix = base.split("实勘")[0]
        category = _FILE_TO_CATEGORY.get(prefix)
        if not category:
            print(f"跳过未识别类别：{base}")
            continue
        out[category] = _extract(path)
        total = sum(len(g["items"]) for g in out[category])  # type: ignore[arg-type]
        print(f"{category}: {total} 因素 / {len(out[category])} 组")

    body = json.dumps(out, ensure_ascii=False, indent=2)
    text = (
        "// 由 tools/gen_survey_factors.py 从 7 份实勘表自动生成，勿手改。\n"
        "// 每类别逐因素清单（区位/实物/权益状况），现场采集页按类别渲染「描述+档次」。\n"
        f"module.exports = {body};\n"
    )
    with open(_OUT, "w", encoding="utf-8") as fh:
        fh.write(text)
    print(f"已写 {_OUT}")


if __name__ == "__main__":
    main()
